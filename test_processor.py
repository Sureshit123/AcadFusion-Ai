import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

def generate_excel_report(results):
    valid_results = [r for r in results if r.get('status') in ['Pass', 'Fail']]
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        if valid_results:
            all_subject_codes = set()
            for r in valid_results:
                all_subject_codes.update(r.get('subjects', {}).keys())
            all_subject_codes = sorted(list(all_subject_codes))
            
            df_all = []
            for i, r in enumerate(valid_results):
                row = {
                    ('Basic', 'SL.No'): i + 1,
                    ('Basic', 'USN'): r.get('usn'),
                    ('Basic', 'NAME'): r.get('name'),
                }
                
                backlog_count = 0
                for sub_code in all_subject_codes:
                    sub_data = r.get('subjects', {}).get(sub_code)
                    if sub_data:
                        row[(sub_code, 'INT')] = sub_data.get('internal', 0)
                        row[(sub_code, 'EXT')] = sub_data.get('external', 0)
                        row[(sub_code, 'TOT')] = sub_data.get('total', 0)
                        res_flag = sub_data.get('result', '')
                        row[(sub_code, 'PT')] = res_flag
                        if res_flag in ['F', 'A', 'FAIL', 'ABSENT']:
                            backlog_count += 1
                row[('Summary', 'Grand Total')] = r.get('total_marks', 0)
                row[('Summary', 'No.of B/L')] = backlog_count
                df_all.append(row)
            
            df = pd.DataFrame(df_all)
            df.columns = pd.MultiIndex.from_tuples(df.columns)
            
            # Use index=True to bypass NotImplementedError, then delete the index column
            df.to_excel(writer, sheet_name='All Students', index=True)
            worksheet = writer.sheets['All Students']
            worksheet.delete_cols(1)

            # Optional: Add openpyxl styling
            yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            header_font = Font(bold=True)
            max_col = worksheet.max_column
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=max_col), start=1):
                pass
            
    output.seek(0)
    return output

results = [{'usn':'123', 'name':'abc', 'status':'Pass', 'subjects': {'CS1': {'internal':10, 'external':20, 'total':30, 'result':'P'}}, 'total_marks':30, 'max_marks':100}]
try:
    generate_excel_report(results)
    print('SUCCESS')
except Exception as e:
    print('ERROR:')
    traceback.print_exc()
