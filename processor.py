import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.drawing.colors import ColorChoice
from openpyxl.utils import get_column_letter
import io
import json

def classify_grade(total, internal=0, external=0):
    """Categorizes marks into grades."""
    try:
        total = int(total)
        internal = int(internal)
        external = int(external)
    except:
        return "Fail"
        
    # Rule: 50 < internal <= 100 is always PASS
    if 50 < internal <= 100:
        pass
    # Rule: internal <= 50 and external < 18 is FAIL
    elif internal <= 50 and external < 18:
        return "Fail"
    
    if total >= 91: return "Outstanding / Excellent"
    if total >= 81: return "Very Good"
    if total >= 71: return "Good"
    if total >= 61: return "Above Average"
    if total >= 51: return "Average"
    if total >= 41: return "Pass"
    return "Fail"

def generate_excel_report(results):
    """
    Generates a multi-sheet Excel file with Overall Summary and Subject-wise Analysis.
    Following ISE Department college format.
    """
    valid_results = [r for r in results if r.get('status') in ['Pass', 'Fail']]
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # --- SHEET 1: OVERALL RESULT SUMMARY ---
        summary_data = {
            'FCD': 0, 'FC': 0, 'SC': 0, 'Fail': 0, 'Absent': 0
        }
        
        for r in results:
            status = r.get('status')
            if status == "Pass":
                total = r.get('total_marks', 0)
                max_m = r.get('max_marks', 1)
                per = (total / max_m) * 100
                if per >= 70: summary_data['FCD'] += 1
                elif per >= 60: summary_data['FC'] += 1
                elif per >= 35: summary_data['SC'] += 1
                else: summary_data['Fail'] += 1
            elif status == "Fail":
                summary_data['Fail'] += 1
            elif "Absent" in status or "No Res" in status:
                summary_data['Absent'] += 1

        total_pass = summary_data['FCD'] + summary_data['FC'] + summary_data['SC']
        appeared = total_pass + summary_data['Fail']
        pass_percentage = round((total_pass / appeared) * 100, 2) if appeared > 0 else 0
        
        df_overall = pd.DataFrame([{
            'FCD': summary_data['FCD'],
            'FC': summary_data['FC'],
            'SC': summary_data['SC'],
            'Fail': summary_data['Fail'],
            'Absent': summary_data['Absent'],
            'Total Pass': total_pass,
            'Passing Percentage': pass_percentage,
            'Total No of Students Appeared': appeared
        }])
        
        df_overall.to_excel(writer, sheet_name='OVERALL RESULT SUMMARY', index=False, startrow=1)
        ws_sum = writer.sheets['OVERALL RESULT SUMMARY']
        
        # Header: Overall Result
        ws_sum.merge_cells('A1:H1')
        ws_sum['A1'] = "Overall Result"
        ws_sum['A1'].font = Font(bold=True, size=14)
        ws_sum['A1'].alignment = Alignment(horizontal='center')

        # Formatting Overall Summary Table
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws_sum.iter_rows(min_row=2, max_row=3, min_col=1, max_col=8):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                if row[0].row == 2: cell.font = Font(bold=True)

        # Footer
        ws_sum.cell(row=7, column=1, value="Coordinator").font = Font(bold=True)
        ws_sum.cell(row=7, column=8, value="HOD").font = Font(bold=True)
        ws_sum.cell(row=7, column=8).alignment = Alignment(horizontal='right')

        # --- SHEET 2: SUBJECT-WISE RESULT ANALYSIS ---
        all_subjects_meta = {} # code -> name
        for r in valid_results:
            for sub_code, sub_data in r.get('subjects', {}).items():
                if sub_code not in all_subjects_meta:
                    all_subjects_meta[sub_code] = sub_data.get('name', 'N/A')

        if valid_results:
            sub_analysis = []
            for sub_code, sub_name in all_subjects_meta.items():
                stats = {'fcd': 0, 'fc': 0, 'sc': 0, 'fail': 0, 'absent': 0, 'appeared': 0}
                for r in valid_results:
                    sub_data = r.get('subjects', {}).get(sub_code)
                    if sub_data:
                        res = sub_data.get('result', '').upper()
                        total = sub_data.get('total', 0)
                        if res in ['A', 'ABSENT']:
                            stats['absent'] += 1
                        else:
                            stats['appeared'] += 1
                            if res in ['P', 'PASS']:
                                if total >= 70: stats['fcd'] += 1
                                elif total >= 60: stats['fc'] += 1
                                elif total >= 35: stats['sc'] += 1
                                else: stats['fail'] += 1
                            else:
                                stats['fail'] += 1
                
                passed = stats['fcd'] + stats['fc'] + stats['sc']
                p_per = round((passed / stats['appeared']) * 100, 2) if stats['appeared'] > 0 else 0
                
                sub_analysis.append({
                    'Subjects': sub_name,
                    'Sub code': sub_code,
                    'FCD (70-100%)': stats['fcd'],
                    'FC (60-69%)': stats['fc'],
                    'SC (35-59%)': stats['sc'],
                    'Fail': stats['fail'],
                    'Absent': stats['absent'],
                    'Total students Appeared': stats['appeared'],
                    'No of student passed': passed,
                    'Passing %age': p_per
                })
            
            df_sub = pd.DataFrame(sub_analysis)
            df_sub.to_excel(writer, sheet_name='SUBJECT-WISE RESULT ANALYSIS', index=False, startrow=3)
            ws_sub = writer.sheets['SUBJECT-WISE RESULT ANALYSIS']
            
            # Title for Sheet 2
            ws_sub.merge_cells('A1:J1')
            ws_sub['A1'] = "SUBJECT-WISE RESULT ANALYSIS"
            ws_sub['A1'].font = Font(bold=True, size=12)
            ws_sub['A1'].alignment = Alignment(horizontal='center')

            # Formatting & Conditional Color
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            for row_idx, row in enumerate(ws_sub.iter_rows(min_row=4, max_row=ws_sub.max_row, min_col=1, max_col=10), 4):
                for cell in row:
                    cell.border = thin_border
                    if row_idx == 4:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', wrap_text=True)
                    else:
                        if cell.column >= 3:
                            cell.alignment = Alignment(horizontal='center')
                        if cell.column == 10:
                            if isinstance(cell.value, (int, float)):
                                cell.fill = green_fill if cell.value >= 80 else red_fill

            # Auto-adjust column widths
            for col in ws_sub.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                ws_sub.column_dimensions[column_letter].width = min(max_length + 2, 40)

            # --- ADD CHART TO SHEET 1 (Referencing Sheet 2 data) ---
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Subject wise %"
            chart.y_axis.title = 'PERCENTAGE'
            chart.x_axis.title = 'SUBJECTS'
            chart.height = 12
            chart.width = 25
            
            # Rotation for X-axis labels (45 degrees)
            chart.x_axis.labelRotation = 4500 

            # Show percentage values on top of bars
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showVal = True
            
            # Data from Sheet 2
            data = Reference(ws_sub, min_col=10, min_row=4, max_row=ws_sub.max_row)
            cats = Reference(ws_sub, min_col=1, min_row=5, max_row=ws_sub.max_row)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.legend = None
            
            # Set bar color (Series 1)
            if chart.series:
                chart.series[0].graphical_properties = GraphicalProperties(solidFill=ColorChoice(srgbClr="4F81BD"))

            # Position on Sheet 1
            ws_sum.add_chart(chart, "A12")

        # --- SHEET 3: All Students ---
        all_subject_codes = sorted(list(all_subjects_meta.keys()))
        df_all = []
        for i, r in enumerate(results):
            res_status = r.get('status')
            row = {'SL.No': i + 1, 'USN': r.get('usn'), 'NAME': r.get('name') if r.get('name') else res_status}
            backlog_count = 0
            for sub_code in all_subject_codes:
                if res_status in ['Pass', 'Fail']:
                    sub_data = r.get('subjects', {}).get(sub_code)
                    if sub_data:
                        row[f'{sub_code}_INT'] = sub_data.get('internal', 0)
                        row[f'{sub_code}_EXT'] = sub_data.get('external', 0)
                        row[f'{sub_code}_TOT'] = sub_data.get('total', 0)
                        res_flag = str(sub_data.get('result', '')).upper()
                        row[f'{sub_code}_PT'] = res_flag
                        if res_flag in ['F', 'A', 'ABSENT']: backlog_count += 1
                    else:
                        for k in ['INT', 'EXT', 'TOT', 'PT']: row[f'{sub_code}_{k}'] = '-'
                else:
                    for k in ['INT', 'EXT', 'TOT', 'PT']: row[f'{sub_code}_{k}'] = '-'
            row['Grand Total'] = r.get('total_marks', 0) if res_status in ['Pass', 'Fail'] else '-'
            row['No.of B/L'] = backlog_count if res_status in ['Pass', 'Fail'] else '-'
            df_all.append(row)
        
        df_all_final = pd.DataFrame(df_all)
        df_all_final.to_excel(writer, sheet_name='All Students', index=False, startrow=1)
        ws_all = writer.sheets['All Students']
        
        # Advanced Merged Headers for Sheet 3
        # Static columns vertical merge
        header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        yellow_row_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        for col_idx, col_name in enumerate(['SL.No', 'USN', 'NAME'], 1):
            ws_all.cell(row=1, column=col_idx, value=col_name)
            ws_all.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

        # Dynamic Subject Headers (Horizontal Merge)
        curr_col = 4
        for sub_code in all_subject_codes:
            sub_name = all_subjects_meta.get(sub_code, "N/A")
            ws_all.cell(row=1, column=curr_col, value=f"{sub_code} - {sub_name}")
            ws_all.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=curr_col+3)
            # Row 2 sub-headers
            for off, txt in enumerate(['INT', 'EXT', 'TOT', 'PT']):
                ws_all.cell(row=2, column=curr_col+off, value=txt)
            curr_col += 4
            
        # Summary columns vertical merge
        for col_name in ['Grand Total', 'No.of B/L']:
            ws_all.cell(row=1, column=curr_col, value=col_name)
            ws_all.merge_cells(start_row=1, start_column=curr_col, end_row=2, end_column=curr_col)
            curr_col += 1

        # Apply Styling & Backlog Highlighting
        for row_idx, row in enumerate(ws_all.iter_rows(min_row=1, max_row=ws_all.max_row, min_col=1, max_col=ws_all.max_column), 1):
            # Header Styling
            if row_idx <= 2:
                for cell in row:
                    cell.font = Font(bold=True, size=9)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border
            else:
                # Backlog Highlighting
                bl_cell = row[-1] # No.of B/L is the last column
                has_backlog = False
                try:
                    if isinstance(bl_cell.value, int) and bl_cell.value > 0: has_backlog = True
                except: pass
                
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                    if has_backlog:
                        cell.fill = yellow_row_fill
        
        # Auto-adjust column widths for Sheet 3
        ws_all.column_dimensions['B'].width = 15 # USN
        ws_all.column_dimensions['C'].width = 30 # NAME

        # --- SHEET 4: Top 5 Toppers ---
        toppers = sorted([r for r in valid_results if r.get('status') == 'Pass'], key=lambda x: x.get('total_marks', 0), reverse=True)[:5]
        df_toppers = pd.DataFrame([{
            'Rank': i+1, 'USN': t['usn'], 'Name': t['name'], 'Total': t['total_marks']
        } for i, t in enumerate(toppers)])
        df_toppers.to_excel(writer, sheet_name='Top 5 Toppers', index=False)

    output.seek(0)
    return output
